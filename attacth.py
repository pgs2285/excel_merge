from select import select
import openpyxl as op
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import sys



def attachStart():
    wp = planFile_wb.active
    usingCell = int()
    usingCellRow = int()
    getter = select_week.get()

    for i in range(1,15): #추후 일정 추출을 위해 column 저장
        for j in range(1,5):
            tmp = wp.cell(row = j, column=i).value
            if str(tmp) == "이용시간":
                usingCellColumn = i
                usingCellRow = j

    for sheet in sheet_list:
        wp = planFile_wb[f'{sheet}']
        if getter =="1주":
            print("firstCell : " + str(firstWeekStart) + " ~ " + str(firstWeekEnd))       
            for i in range(usingCellRow+1, 100):
                time = str(wp.cell(row = i, column= usingCellColumn).value)
                if time != 'None' and time [:5] != "00:00" and time[:5] != "05:00": #wp(주간 방문계획에서 시간을 가져옴)
                    for j in range(firstWeekStart, firstWeekEnd+1):
                        plan = str(ws.cell(row= j , column= timeCell).value) # 주간방문계획에서 가져온 시간을 일정표와 비교하기 위해 가져옴
                        # print(plan[:5], time[:5])
                        if plan[:5] == time[:5]: #만약 가져온 시간과 일정표에 존재하는 시간이 매치하면
                            print(plan[:5], time[:5])  
                            for k in range(0,15):
                                if ws.cell(row= j , column= planCell + k).value != None:
                                    ans = str(ws.cell(row=j, column = planCell + k).value)
                                    wp.cell(row = i, column = usingCellColumn+k+1, value = ans)

        elif getter =="2주":
            print("secondCell : " + str(secondWeekStart) + " ~ " + str(secondWeekEnd))
            for i in range(usingCellRow+1, 100):
                time = str(wp.cell(row = i, column= usingCellColumn).value)
                if time != 'None' and time [:5] != "00:00" and time[:5] != "05:00": #wp(주간 방문계획에서 시간을 가져옴)
                    for j in range(secondWeekStart, secondWeekEnd+1):
                        plan = str(ws.cell(row= j , column= timeCell).value) # 주간방문계획에서 가져온 시간을 일정표와 비교하기 위해 가져옴
                        # print(plan[:5], time[:5])
                        if plan[:5] == time[:5]: #만약 가져온 시간과 일정표에 존재하는 시간이 매치하면
                            print(plan[:5], time[:5])  
                            for k in range(0,15):
                                if ws.cell(row= j , column= planCell + k).value != None:
                                    ans = str(ws.cell(row=j, column = planCell + k).value)
                                    wp.cell(row = i, column = usingCellColumn+k+1, value = ans)

        elif getter == "3주":
            print("thirdCell : " + str(thirdWeekStart) + " ~ " + str(thirdWeekEnd))
            for i in range(usingCellRow+1, 100):
                time = str(wp.cell(row = i, column= usingCellColumn).value)
                if time != 'None' and time [:5] != "00:00" and time[:5] != "05:00": #wp(주간 방문계획에서 시간을 가져옴)
                    for j in range(thirdWeekStart, thirdWeekEnd+1):
                        plan = str(ws.cell(row= j , column= timeCell).value) # 주간방문계획에서 가져온 시간을 일정표와 비교하기 위해 가져옴
                        # print(plan[:5], time[:5])
                        if plan[:5] == time[:5]: #만약 가져온 시간과 일정표에 존재하는 시간이 매치하면
                            print(plan[:5], time[:5])  
                            for k in range(0,15):
                                if ws.cell(row= j , column= planCell + k).value != None:
                                    ans = str(ws.cell(row=j, column = planCell + k).value)
                                    try:
                                        print(ans)
                                        ans = int(ans)
                                    except:
                                        pass    
                                    wp.cell(row = i, column = usingCellColumn+k+1, value = ans)

        else:
            messagebox.showerror("주차 선택", "주차를 선택을 해주세요")
        planFile_wb.save(filename = "./test.xlsx")    

        


root = Tk() #객체 인스턴스 생성
tmp = []
tmp2 = []
root.title("build") # 타이틀 설정
root.geometry("200x200") # 크기 설정
root.resizable(False,False) # 크기변경 불가

list_file = []                                          #파일 목록 담을 리스트 생성

week_file = filedialog.askopenfilename(initialdir="/",\
                 title = "주간 일정 파일을 선택 해 주세요",\
                    filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls"),("*.csv","*csv")))

plan_file = filedialog.askopenfilename(initialdir="/",\
                 title = "서비스제공표 파일을 선택 해 주세요",\
                    filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls"),("*.csv","*csv")))


weekFile_wb = op.load_workbook(r""+plan_file,data_only=True)
print('sheet count: ', len(weekFile_wb.sheetnames))
ws = weekFile_wb.active

timeCell = int()
planCell = int()

planFile_wb = op.load_workbook(r""+week_file,data_only=True)
wp = planFile_wb.active
sheet_list = planFile_wb.sheetnames
for i in range(1,10): #추후 일정 추출을 위해 column 저장
    tmp = ws.cell(row = 1, column=i).value
    
    if tmp == '시간':
        timeCell = i
    elif tmp == '일정': 
        planCell = i   

cnt_1 = 0
cnt_2 = 0
cnt_3 = 0

firstWeekStart = int()
firstWeekEnd = int()
secondWeekStart = int()
secondWeekEnd = int() 
thirdWeekStart = int()
thirdWeekEnd = int()    
for i in range(1,100):
    tmp = ws.cell(row = i, column = 1).value

    if tmp == '1주차':
        firstWeekEnd = i
        cnt_1 += 1
    elif tmp == '2주차':
        secondWeekEnd = i
        cnt_2 += 1
    elif tmp == '3주차':
        thirdWeekEnd = i
        cnt_3 += 1

firstWeekStart = firstWeekEnd - cnt_1
secondWeekStart = secondWeekEnd - cnt_2
thirdWeekStart = thirdWeekEnd - cnt_3




select_week = ttk.Combobox(root, value=['1주','2주','3주'])
select_week.place(x = 10, y = 80)
text1 = Label(root, text = "주차를 선택해 주세요")
text1.place(x=10, y=30)
startButton = Button(root, text='작업 시작하기',command=attachStart)
startButton.place(x=10,y=120)
root.mainloop()

